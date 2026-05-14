from flask import Flask, request, jsonify, render_template, send_file
import google.generativeai as genai
import os
import json
import sqlite3
from datetime import datetime
from werkzeug.utils import secure_filename
import PyPDF2
from docx import Document
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    filename='app.log',
                    filemode='a')
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'docx', 'txt'}

# Configure Gemini AI
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
if not GEMINI_API_KEY:
    print("WARNING: GEMINI_API_KEY not found in environment variables!")
else:
    genai.configure(api_key=GEMINI_API_KEY)

# Create uploads folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Database initialization
def init_db():
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Create table if it doesn't exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requirement_file TEXT NOT NULL,
            requirement_id TEXT,
            test_case_name TEXT NOT NULL,
            description TEXT,
            preconditions TEXT,
            test_steps TEXT,
            expected_result TEXT,
            priority TEXT,
            test_type TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Check if requirement_id column exists (migration for existing DB)
    cursor = conn.execute("PRAGMA table_info(test_cases)")
    columns = [column[1] for column in cursor.fetchall()]
    if 'requirement_id' not in columns:
        print("Migrating database: Adding requirement_id column...")
        conn.execute('ALTER TABLE test_cases ADD COLUMN requirement_id TEXT')
    
    if 'requirement_description' not in columns:
        print("Migrating database: Adding requirement_description column...")
        conn.execute('ALTER TABLE test_cases ADD COLUMN requirement_description TEXT')
    
    if 'status' not in columns:
        print("Migrating database: Adding status column...")
        conn.execute('ALTER TABLE test_cases ADD COLUMN status TEXT DEFAULT "Not Executed"')
    
    conn.commit()
    conn.close()
# Initialize database on startup
init_db()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def extract_text_from_pdf(file_path):
    """Extract text from PDF file"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
    except Exception as e:
        print(f"Error extracting PDF: {e}")
        return None

def extract_text_from_docx(file_path):
    """Extract text from DOCX file"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Error extracting DOCX: {e}")
        return None

def extract_text_from_txt(file_path):
    """Extract text from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except Exception as e:
        print(f"Error extracting TXT: {e}")
        return None

def extract_text_from_file(file_path, file_extension):
    """Extract text based on file type"""
    if file_extension == 'pdf':
        return extract_text_from_pdf(file_path)
    elif file_extension == 'docx':
        return extract_text_from_docx(file_path)
    elif file_extension == 'txt':
        return extract_text_from_txt(file_path)
    return None

def generate_test_cases_with_gemini(requirement_text):
    """Use Gemini AI to generate test cases from requirements"""
    try:
        # Use specific model version
        # gemini-flash-latest seems to map to 2.5-flash which has strict quotas but is available
        model = genai.GenerativeModel('gemini-flash-latest')
        
        prompt = f"""
        You are an expert QA engineer with 10+ years of experience. Analyze the following requirement document THOROUGHLY.
        First, identifying ALL distinct requirements and assign them IDs (e.g., REQ-001, REQ-002) if they are not already numbered.
        Then, generate COMPREHENSIVE test cases covering each of these requirements.

        REQUIREMENT DOCUMENT:
        {requirement_text}

        CRITICAL INSTRUCTIONS:
        1. READ THE ENTIRE DOCUMENT carefully and identify EVERY feature, function, and requirement
        2. For EACH feature/requirement, generate test cases covering:
           - POSITIVE scenarios (happy path)
           - NEGATIVE scenarios (error handling)
           - EDGE CASES (boundary conditions)
           - INTEGRATION scenarios

        3. Generate detailed test cases in JSON format. Each test case MUST include:
           - test_case_name: A clear, specific, descriptive name
           - requirement_id: The ID of the requirement this test case covers (e.g., "REQ-001")
           - requirement_description: A brief description of the requirement itself (e.g. "User must log in")
           - description: Brief description of what is being tested and why
           - preconditions: Any setup or conditions needed before testing
           - test_steps: Detailed step-by-step instructions (as a numbered list with \\n separators)
           - expected_result: What should happen if the test passes
           - priority: High, Medium, or Low
           - test_type: Functional, Security, Performance, Usability, etc.

        TEST COVERAGE REQUIREMENTS:
        - Ensure EVERY identified requirement has at least one Positive and one Negative test case.
        - Maintain a 50-50 balance between Functional and Non-Functional tests where applicable.

        Return ONLY a valid JSON array of test cases.

        Example format:
        [
          {{
            "test_case_name": "Verify successful user login",
            "requirement_id": "REQ-001",
            "requirement_description": "The system shall allow users to log in with valid credentials",
            "description": "Test that users can successfully log in using correct email and password",
            "preconditions": "1. User account exists\\n2. User is not logged in",
            "test_steps": "1. Navigate to login page\\n2. Enter valid email\\n3. Enter valid password\\n4. Click Login",
            "expected_result": "User is successfully authenticated and redirected to dashboard",
            "priority": "High",
            "test_type": "Functional"
          }}
        ]
        """
        
        print(f"Starting test case generation...")
        print(f"Requirement text length: {len(requirement_text)} characters")
        
        response = model.generate_content(
            prompt,
            generation_config={"response_mime_type": "application/json"}
        )
        
        print("Received response from Gemini API")
        response_text = response.text.strip()
        print(f"Response length: {len(response_text)} characters")
        print(f"Raw response start: {response_text[:100]}...")
        
        # Clean up markdown code blocks if present
        if response_text.startswith('```json'):
            response_text = response_text[7:]
        if response_text.startswith('```'):
            response_text = response_text[3:]
        if response_text.endswith('```'):
            response_text = response_text[:-3]
            
        response_text = response_text.strip()
        
        # Parse JSON response
        print("Parsing JSON response...")
        test_cases = json.loads(response_text)
        print(f"Successfully parsed {len(test_cases)} test cases")
        return test_cases, None

    except AttributeError as e:
        error_msg = f"API response error: {str(e)}"
        logger.error(error_msg)
        logger.error("This usually means the API key is invalid or the API returned an unexpected response")
        if 'response' in locals():
            logger.error(f"Response object: {response}")
        return None, error_msg

    except Exception as e:
        error_type = type(e).__name__
        error_msg = str(e)
        logger.error(f"Unexpected error generating test cases: {error_type}: {error_msg}")
        
        if "429" in error_msg:
            return None, "Quota exceeded. Please check your Gemini API plan or try again later."
            
        import traceback
        logger.error(f"Full traceback:\n{traceback.format_exc()}")
        return None, f"Error: {error_type}: {error_msg}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and generate test cases"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only PDF, DOCX, and TXT files are allowed'}), 400
    
    try:
        # Save file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Extract text from file
        file_extension = filename.rsplit('.', 1)[1].lower()
        requirement_text = extract_text_from_file(file_path, file_extension)
        
        if not requirement_text:
            return jsonify({'error': 'Failed to extract text from file'}), 500
        
        # Generate test cases using Gemini AI
        print(f"Generating test cases for file: {filename}")
        test_cases, error = generate_test_cases_with_gemini(requirement_text)
        
        if error:
            print(f"ERROR: {error}")
            return jsonify({'error': error}), 500
            
        if not test_cases:
            error_msg = 'Failed to generate test cases. Please check the server logs for details. Common issues: API key invalid, network connectivity, or document format issues.'
            print(f"ERROR: {error_msg}")
            return jsonify({'error': error_msg}), 500
        
        if not isinstance(test_cases, list):
            error_msg = f'Invalid response format from AI. Expected a list, got {type(test_cases).__name__}'
            print(f"ERROR: {error_msg}")
            return jsonify({'error': error_msg}), 500
        
        # Connect to database
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # DELETE PREVIOUS TEST CASES FOR THIS FILE
        cursor.execute('DELETE FROM test_cases WHERE requirement_file = ?', (filename,))
        deleted_count = cursor.rowcount
        
        if deleted_count > 0:
            print(f"Deleted {deleted_count} previous test cases for file: {filename}")
        
        # Save new test cases to database
        saved_test_cases = []
        for tc in test_cases:
            # Handle list fields if Gemini returns them as arrays
            test_steps = tc.get('test_steps', '')
            if isinstance(test_steps, list):
                test_steps = '\n'.join(test_steps)
                
            preconditions = tc.get('preconditions', '')
            if isinstance(preconditions, list):
                preconditions = '\n'.join(preconditions)

            cursor.execute('''
                INSERT INTO test_cases 
                (requirement_file, requirement_id, requirement_description, test_case_name, description, preconditions, test_steps, expected_result, priority, test_type, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                filename,
                tc.get('requirement_id', 'General'),
                tc.get('requirement_description', ''),
                tc.get('test_case_name', ''),
                tc.get('description', ''),
                preconditions,
                test_steps,
                tc.get('expected_result', ''),
                tc.get('priority', 'Medium'),
                tc.get('test_type', 'Functional'),
                'Not Executed'
            ))
            
            test_case_id = cursor.lastrowid
            saved_test_cases.append({
                'id': test_case_id,
                'requirement_file': filename,
                **tc
            })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Successfully generated {len(saved_test_cases)} test cases for {filename}',
            'filename': filename,
            'test_cases': saved_test_cases,
            'replaced': deleted_count > 0
        }), 200
    
    except Exception as e:
        print(f"Error processing file: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases', methods=['GET'])
def get_test_cases():
    """Get all test cases or filter by filename"""
    try:
        # Get optional filename parameter
        filename = request.args.get('filename', None)
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Filter by filename if provided
        if filename:
            cursor.execute('SELECT * FROM test_cases WHERE requirement_file = ? ORDER BY created_at DESC', (filename,))
        else:
            cursor.execute('SELECT * FROM test_cases ORDER BY created_at DESC')
        rows = cursor.fetchall()
        
        test_cases = []
        for row in rows:
            # Handle possible missing column for old records or if migration failed silently (though we added it)
            req_id = row['requirement_id'] if 'requirement_id' in row.keys() else 'General'
            
            test_cases.append({
                'id': row['id'],
                'requirement_file': row['requirement_file'],
                'requirement_id': req_id,
                'requirement_description': row['requirement_description'] if 'requirement_description' in row.keys() else '',
                'test_case_name': row['test_case_name'],
                'description': row['description'],
                'preconditions': row['preconditions'],
                'test_steps': row['test_steps'],
                'expected_result': row['expected_result'],
                'priority': row['priority'],
                'test_type': row['test_type'],
                'status': row['status'] if 'status' in row.keys() else 'Not Executed',
                'created_at': row['created_at']
            })
        
        conn.close()
        return jsonify(test_cases), 200
    
    except Exception as e:
        print(f"Error fetching test cases: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases/<int:test_case_id>', methods=['PUT'])
def update_test_case(test_case_id):
    """Update a test case"""
    try:
        data = request.json
        
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE test_cases 
            SET test_case_name = ?, description = ?, preconditions = ?, 
                test_steps = ?, expected_result = ?, priority = ?, test_type = ?, requirement_description = ?, status = ?
            WHERE id = ?
        ''', (
            data.get('test_case_name'),
            data.get('description'),
            data.get('preconditions'),
            data.get('test_steps'),
            data.get('expected_result'),
            data.get('priority'),
            data.get('test_type'),
            data.get('requirement_description'),
            data.get('status'),
            test_case_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Test case updated successfully'}), 200
    
    except Exception as e:
        print(f"Error updating test case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases/<int:test_case_id>', methods=['DELETE'])
def delete_test_case(test_case_id):
    """Delete a test case"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM test_cases WHERE id = ?', (test_case_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Test case deleted successfully'}), 200
    
    except Exception as e:
        print(f"Error deleting test case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/generate-automation', methods=['POST'])
def generate_automation_script():
    try:
        print("--- Automation Request Received ---")
        # Check if we have a file or just JSON
        test_case_data = request.form.get('test_case')
        image_file = request.files.get('screenshot')
        
        print(f"Test Case Data Present: {bool(test_case_data)}")
        print(f"Image File Present: {bool(image_file)}")
        
        if not test_case_data:
            print("Error: No test case data")
            return jsonify({'error': 'No test case data provided'}), 400
            
        import json
        test_case = json.loads(test_case_data)
        print(f"Test Case Name: {test_case.get('test_case_name')}")
        
        # Prepare the prompt
        prompt_text = f"""
        You are an expert QA Automation Engineer. 
        Generate a Playwright test script in JavaScript for the following test case.
        
        Test Case Name: {test_case.get('test_case_name')}
        Description: {test_case.get('description')}
        Preconditions: {test_case.get('preconditions')}
        Steps: {test_case.get('test_steps')}
        Expected Result: {test_case.get('expected_result')}
        
        Rules:
        1. Use 'require("@playwright/test")' syntax.
        2. Use robust USER-FACING locators (getByRole, getByLabel, getByPlaceholder, getByText).
        3. AVOID using IDs or CSS selectors unless absolutely necessary.
        4. Include comments explaining the steps.
        5. Assume a generic base URL 'http://localhost:3000' if not specified.
        6. Return ONLY the raw JavaScript code. Do not include markdown formatting.
        7. Include assertions based on the Expected Result.
        """
        
        model = genai.GenerativeModel('gemini-flash-latest')
        
        if image_file:
            print(f"Processing screenshot: {image_file.filename}")
            # If screenshot provided, use it!
            from PIL import Image
            image = Image.open(image_file)
            print(f"Image opened successfully. Size: {image.size}")
            
            prompt_text += "\n\nI have attached a screenshot of the page. Use this visual information to determine the most accurate locators (button names, labels, placeholders)."
            
            print("Sending request with image to Gemini...")
            response = model.generate_content([prompt_text, image])
        else:
            print("Sending text-only request to Gemini...")
            # Text only generation
            response = model.generate_content(prompt_text)
        
        print("Response received from Gemini")
        # Clean up response
        script_content = response.text.replace('```javascript', '').replace('```', '').strip()
        print(f"Generated script length: {len(script_content)}")
        
        return jsonify({'script': script_content})

    except Exception as e:
        print(f"Error generating automation script: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/test-ai', methods=['GET'])
def test_ai_connection():
    try:
        model = genai.GenerativeModel('gemini-1.5-flash')
        response = model.generate_content("Reply with 'API IS WORKING'")
        return jsonify({'status': 'success', 'message': response.text})
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/test-cases/clear-all', methods=['DELETE'])
def clear_all_test_cases():
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM test_cases')
        deleted_count = cursor.rowcount
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': f'Successfully deleted {deleted_count} test cases'}), 200
    
    except Exception as e:
        print(f"Error clearing test cases: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export', methods=['GET'])
def export_test_cases():
    try:
        print("Export request received")
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM test_cases ORDER BY created_at DESC')
        rows = cursor.fetchall()
        
        conn.close()
        
        print(f"Found {len(rows)} test cases to export")
        
        if len(rows) == 0:
            return jsonify({'error': 'No test cases to export'}), 400
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define headers
        headers = ['ID', 'Test Case Name', 'Description', 'Preconditions', 
                   'Test Steps', 'Expected Result', 'Priority', 'Test Type', 'Status',
                   'Requirement File', 'Created At']
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for row_num, row in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=row['id'])
            ws.cell(row=row_num, column=2, value=row['test_case_name'])
            ws.cell(row=row_num, column=3, value=row['description'])
            ws.cell(row=row_num, column=4, value=row['preconditions'])
            ws.cell(row=row_num, column=5, value=row['test_steps'])
            ws.cell(row=row_num, column=6, value=row['expected_result'])
            ws.cell(row=row_num, column=7, value=row['priority'])
            ws.cell(row=row_num, column=8, value=row['test_type'])
            ws.cell(row=row_num, column=9, value=row['status'] if 'status' in row.keys() else 'Not Executed')
            ws.cell(row=row_num, column=10, value=row['requirement_file'])
            ws.cell(row=row_num, column=11, value=row['created_at'])
            
            # Apply text wrapping for long content
            for col in [2, 3, 4, 5, 6]:
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 20
        
        # Freeze the header row
        ws.freeze_panes = 'A2'

        # ==========================================
        # Create Traceability Matrix Sheet
        # ==========================================
        ws_rtm = wb.create_sheet(title="Traceability Matrix")
        
        # Define RTM headers
        rtm_headers = ['Requirement ID', 'Requirement Description', 'Test Case ID', 'Test Case Name', 'Status']
        
        # Write RTM headers
        for col_num, header in enumerate(rtm_headers, 1):
            cell = ws_rtm.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Group test cases by Requirement ID
        from collections import defaultdict
        req_map = defaultdict(list)
        for row in rows:
            # Handle potential missing requirement_id key if row structure varies (though we used select *)
            # Also handle if requirement_id is None
            r_id = row['requirement_id'] if row['requirement_id'] else 'General'
            req_map[r_id].append(row)
            
        # Sort requirements
        sorted_req_ids = sorted(req_map.keys(), key=lambda x: (x == 'General', x))

        current_row = 2
        
        # Styles for RTM
        center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        for req_id in sorted_req_ids:
            cases = req_map[req_id]
            start_row = current_row
            
            for tc in cases:
                # Requirement ID (will be merged later)
                ws_rtm.cell(row=current_row, column=1, value=req_id).alignment = center_align
                
                # Requirement Description (will be merged later)
                req_desc = tc['requirement_description'] if 'requirement_description' in tc.keys() and tc['requirement_description'] else 'No description'
                ws_rtm.cell(row=current_row, column=2, value=req_desc).alignment = left_align
                
                # Test Case ID
                ws_rtm.cell(row=current_row, column=3, value=f"TC-{tc['id']}").alignment = center_align
                
                # Test Case Name
                ws_rtm.cell(row=current_row, column=4, value=tc['test_case_name']).alignment = left_align
                
                # Status
                status = tc['status'] if 'status' in tc.keys() and tc['status'] else 'Not Executed'
                ws_rtm.cell(row=current_row, column=5, value=status).alignment = center_align
                
                current_row += 1
            
            end_row = current_row - 1
            
            # Merge Requirement ID and Description cells if there are multiple test cases
            if end_row > start_row:
                ws_rtm.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws_rtm.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

        # Adjust RTM column widths
        ws_rtm.column_dimensions['A'].width = 20
        ws_rtm.column_dimensions['B'].width = 50
        ws_rtm.column_dimensions['C'].width = 15
        ws_rtm.column_dimensions['D'].width = 40
        ws_rtm.column_dimensions['E'].width = 15

        # Freeze RTM header
        ws_rtm.freeze_panes = 'A2'
        
        # Save to BytesIO instead of file
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        export_filename = f'test_cases_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        print(f"Sending file: {export_filename}")
        
        # Return file from memory
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=export_filename
        )
    
    except Exception as e:
        print(f"Error exporting test cases: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
