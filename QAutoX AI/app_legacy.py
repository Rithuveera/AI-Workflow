from flask import Flask, request, jsonify, render_template, send_file
import google.generativeai as genai
import os
import json
import sqlite3
from datetime import datetime
from werkzeug.utils import secure_filename
import PyPDF2
from docx import Document
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    filename='app.log',
                    filemode='a')
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'docx', 'txt'}

# Configure Gemini AI
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
if not GEMINI_API_KEY:
    print("WARNING: GEMINI_API_KEY not found in environment variables!")
else:
    genai.configure(api_key=GEMINI_API_KEY)

# Create uploads folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Database initialization
def init_db():
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    
    # Create table if it doesn't exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requirement_file TEXT NOT NULL,
            requirement_id TEXT,
            test_case_name TEXT NOT NULL,
            description TEXT,
            preconditions TEXT,
            test_steps TEXT,
            expected_result TEXT,
            priority TEXT,
            test_type TEXT,
            test_suite TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Check if requirement_id column exists (migration for existing DB)
    cursor = conn.execute("PRAGMA table_info(test_cases)")
    columns = [column[1] for column in cursor.fetchall()]
    if 'requirement_id' not in columns:
        print("Migrating database: Adding requirement_id column...")
        conn.execute('ALTER TABLE test_cases ADD COLUMN requirement_id TEXT')
    
    if 'requirement_description' not in columns:
        print("Migrating database: Adding requirement_description column...")
        conn.execute('ALTER TABLE test_cases ADD COLUMN requirement_description TEXT')
    
    if 'status' not in columns:
        print("Migrating database: Adding status column...")
        conn.execute('ALTER TABLE test_cases ADD COLUMN status TEXT DEFAULT "Not Executed"')
    
    if 'test_suite' not in columns:
        print("Migrating database: Adding test_suite column...")
        conn.execute('ALTER TABLE test_cases ADD COLUMN test_suite TEXT')
        
    # Create executions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_executions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            test_case_id INTEGER,
            version TEXT,
            status TEXT,
            executed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            comments TEXT,
            defect_id TEXT,
            evidence_file TEXT,
            FOREIGN KEY (test_case_id) REFERENCES test_cases (id)
        )
    ''')
    
    # Migration for new columns in test_executions (if table already existed without them)
    cursor = conn.execute("PRAGMA table_info(test_executions)")
    exec_columns = [column[1] for column in cursor.fetchall()]
    
    if 'defect_id' not in exec_columns:
        print("Migrating: Adding defect_id to test_executions")
        conn.execute('ALTER TABLE test_executions ADD COLUMN defect_id TEXT')
        
    if 'evidence_file' not in exec_columns:
        print("Migrating: Adding evidence_file to test_executions")
        conn.execute('ALTER TABLE test_executions ADD COLUMN evidence_file TEXT')
    
    conn.commit()
    conn.close()
# Initialize database on startup
init_db()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def extract_text_from_pdf(file_path):
    """Extract text from PDF file"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
    except Exception as e:
        print(f"Error extracting PDF: {e}")
        return None

def extract_text_from_docx(file_path):
    """Extract text from DOCX file"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Error extracting DOCX: {e}")
        return None

def extract_text_from_txt(file_path):
    """Extract text from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except Exception as e:
        print(f"Error extracting TXT: {e}")
        return None

def extract_text_from_file(file_path, file_extension):
    """Extract text based on file type"""
    if file_extension == 'pdf':
        return extract_text_from_pdf(file_path)
    elif file_extension == 'docx':
        return extract_text_from_docx(file_path)
    elif file_extension == 'txt':
        return extract_text_from_txt(file_path)
    return None


def generate_test_cases_with_gemini(requirement_text, model_name='gemini-2.5-flash'):
    """Use Gemini AI to generate test cases from requirements"""
    try:
        # Use specific model version
        print(f"Using Gemini Model: {model_name}")
        model = genai.GenerativeModel(model_name)
        
        prompt = f"""
        You are an expert QA engineer with 10+ years of experience. Analyze the following requirement document THOROUGHLY.
        First, identifying ALL distinct requirements and assign them IDs (e.g., REQ-001, REQ-002) if they are not already numbered.
        Then, generate COMPREHENSIVE test cases covering each of these requirements.

        REQUIREMENT DOCUMENT:
        {requirement_text}

        CRITICAL INSTRUCTIONS:
        1. READ THE ENTIRE DOCUMENT carefully and identify EVERY feature, function, and requirement
        2. For EACH feature/requirement, generate test cases covering:
           - POSITIVE scenarios (happy path)
           - NEGATIVE scenarios (error handling)
           - EDGE CASES (boundary conditions)
           - INTEGRATION scenarios

        3. Generate detailed test cases in JSON format. Each test case MUST include:
           - test_case_name: A clear, specific, descriptive name
           - requirement_id: The ID of the requirement this test case covers (e.g., "REQ-001")
           - requirement_description: A brief description of the requirement itself (e.g. "User must log in")
           - description: Brief description of what is being tested and why
           - preconditions: Any setup or conditions needed before testing
           - test_steps: Detailed step-by-step instructions (as a numbered list with \\n separators)
           - expected_result: What should happen if the test passes
           - priority: High, Medium, or Low
           - test_type: Functional, Security, Performance, Usability, etc.

        TEST COVERAGE REQUIREMENTS:
        - Ensure EVERY identified requirement has at least one Positive and one Negative test case.
        - Maintain a 50-50 balance between Functional and Non-Functional tests where applicable.

        Return ONLY a valid JSON array of test cases.

        Example format:
        [
          {{
            "test_case_name": "Verify successful user login",
            "requirement_id": "REQ-001",
            "requirement_description": "The system shall allow users to log in with valid credentials",
            "description": "Test that users can successfully log in using correct email and password",
            "preconditions": "1. User account exists\\n2. User is not logged in",
            "test_steps": "1. Navigate to login page\\n2. Enter valid email\\n3. Enter valid password\\n4. Click Login",
            "expected_result": "User is successfully authenticated and redirected to dashboard",
            "priority": "High",
            "test_type": "Functional"
          }}
        ]
        """
        
        print(f"Starting test case generation...")
        print(f"Requirement text length: {len(requirement_text)} characters")
        
        response = model.generate_content(
            prompt,
            generation_config={"response_mime_type": "application/json"}
        )
        
        print("Received response from Gemini API")
        response_text = response.text.strip()
        print(f"Response length: {len(response_text)} characters")
        # print(f"Raw response start: {response_text[:100]}...")
        
        # Clean up markdown code blocks if present
        if response_text.startswith('```json'):
            response_text = response_text[7:]
        if response_text.startswith('```'):
            response_text = response_text[3:]
        if response_text.endswith('```'):
            response_text = response_text[:-3]
            
        response_text = response_text.strip()
        
        # Parse JSON response
        print("Parsing JSON response...")
        test_cases = json.loads(response_text)
        print(f"Successfully parsed {len(test_cases)} test cases")
        return test_cases, None

    except AttributeError as e:
        error_msg = f"API response error: {str(e)}"
        logger.error(error_msg)
        logger.error("This usually means the API key is invalid or the API returned an unexpected response")
        if 'response' in locals():
            logger.error(f"Response object: {response}")
        return None, error_msg

    except Exception as e:
        error_type = type(e).__name__
        error_msg = str(e)
        logger.error(f"Unexpected error generating test cases: {error_type}: {error_msg}")
        
        if "429" in error_msg:
            return None, "Quota exceeded. Please try a different model (gemini-2.5-flash usually works best)."
            
        import traceback
        logger.error(f"Full traceback:\n{traceback.format_exc()}")
        return None, f"Error: {error_type}: {error_msg}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and generate test cases"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only PDF, DOCX, and TXT files are allowed'}), 400
    
    try:
        # Save file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        test_suite = request.form.get('test_suite', '')
        model_name = request.form.get('model', 'gemini-2.5-flash')
        
        # Extract text from file
        file_extension = filename.rsplit('.', 1)[1].lower()
        requirement_text = extract_text_from_file(file_path, file_extension)
        
        # Store requirement text for chat context (simple session storage)
        # In production, use Redis or DB. Here global dict keyed by filename is enough for single user
        # or just overwrite a global variable since single user
        global current_context
        current_context = requirement_text
        
        if not requirement_text:
            return jsonify({'error': 'Failed to extract text from file'}), 500
        
        # Generate test cases using Gemini AI
        print(f"Generating test cases for file: {filename} with model {model_name}")
        # Using new fallback function defined at end of file
        test_cases, error = generate_test_cases_fallback(requirement_text, model_name)
        
        if error:
            print(f"ERROR: {error}")
            return jsonify({'error': error}), 500
            
        if not test_cases:
            error_msg = 'Failed to generate test cases. Please check the server logs for details. Common issues: API key invalid, network connectivity, or document format issues.'
            print(f"ERROR: {error_msg}")
            return jsonify({'error': error_msg}), 500
        
        if not isinstance(test_cases, list):
            error_msg = f'Invalid response format from AI. Expected a list, got {type(test_cases).__name__}'
            print(f"ERROR: {error_msg}")
            return jsonify({'error': error_msg}), 500
        
        # Connect to database
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # DELETE PREVIOUS TEST CASES FOR THIS FILE
        cursor.execute('DELETE FROM test_cases WHERE requirement_file = ?', (filename,))
        deleted_count = cursor.rowcount
        
        if deleted_count > 0:
            print(f"Deleted {deleted_count} previous test cases for file: {filename}")
        
        # Save new test cases to database
        saved_test_cases = []
        for tc in test_cases:
            # Handle list fields if Gemini returns them as arrays
            test_steps = tc.get('test_steps', '')
            if isinstance(test_steps, list):
                test_steps = '\n'.join(test_steps)
                
            preconditions = tc.get('preconditions', '')
            if isinstance(preconditions, list):
                preconditions = '\n'.join(preconditions)

            cursor.execute('''
                INSERT INTO test_cases 
                (requirement_file, requirement_id, requirement_description, test_case_name, description, preconditions, test_steps, expected_result, priority, test_type, status, test_suite)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                filename,
                tc.get('requirement_id', 'General'),
                tc.get('requirement_description', ''),
                tc.get('test_case_name', ''),
                tc.get('description', ''),
                preconditions,
                test_steps,
                tc.get('expected_result', ''),
                tc.get('priority', 'Medium'),
                tc.get('test_type', 'Functional'),
                'Not Executed',
                test_suite
            ))
            
            test_case_id = cursor.lastrowid
            saved_test_cases.append({
                'id': test_case_id,
                'requirement_file': filename,
                **tc
            })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Successfully generated {len(saved_test_cases)} test cases for {filename}',
            'filename': filename,
            'test_cases': saved_test_cases,
            'replaced': deleted_count > 0
        }), 200
    
    except Exception as e:
        print(f"Error processing file: {e}")
        return jsonify({'error': str(e)}), 500

# Global context for chat (simple in-memory storage)
current_context = ""

@app.route('/chat', methods=['POST'])
def chat_with_ai():
    """Chat with the AI about the requirements"""
    try:
        data = request.json
        message = data.get('message')
        
        if not message:
            return jsonify({'error': 'No message provided'}), 400
            
        if not current_context:
            return jsonify({'reply': "I don't have any requirement document loaded yet. Please upload a file first."})
            
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        chat_prompt = f"""
        You are a helpful QA Assistant. You have access to the following requirement document:
        
        --- START OF DOCUMENT ---
        {current_context[:30000]} 
        --- END OF DOCUMENT ---
        (Note: Document text may be truncated if too long)
        
        User Query: {message}
        
        Answer the user's question based strictly on the document provided. If the information is not in the document, say so.
        Keep answers concise and helpful.
        """
        
        response = model.generate_content(chat_prompt)
        return jsonify({'reply': response.text})
        
    except Exception as e:
        print(f"Chat error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases', methods=['GET'])
def get_test_cases():
    """Get all test cases or filter by filename"""
    try:
        # Get optional filename parameter
        filename = request.args.get('filename', None)
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Filter by filename if provided
        if filename:
            cursor.execute('SELECT * FROM test_cases WHERE requirement_file = ? ORDER BY created_at DESC', (filename,))
        else:
            cursor.execute('SELECT * FROM test_cases ORDER BY created_at DESC')
        rows = cursor.fetchall()
        
        test_cases = []
        for row in rows:
            # Handle possible missing column for old records or if migration failed silently (though we added it)
            req_id = row['requirement_id'] if 'requirement_id' in row.keys() else 'General'
            
            test_cases.append({
                'id': row['id'],
                'requirement_file': row['requirement_file'],
                'requirement_id': req_id,
                'requirement_description': row['requirement_description'] if 'requirement_description' in row.keys() else '',
                'test_case_name': row['test_case_name'],
                'description': row['description'],
                'preconditions': row['preconditions'],
                'test_steps': row['test_steps'],
                'expected_result': row['expected_result'],
                'priority': row['priority'],
                'test_type': row['test_type'],
                'status': row['status'] if 'status' in row.keys() else 'Not Executed',
                'test_suite': row['test_suite'] if 'test_suite' in row.keys() else '',
                'created_at': row['created_at']
            })
        
        conn.close()
        return jsonify(test_cases), 200
    
    except Exception as e:
        print(f"Error fetching test cases: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases/<int:test_case_id>', methods=['PUT'])
def update_test_case(test_case_id):
    """Update a test case"""
    try:
        data = request.json
        
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE test_cases 
            SET test_case_name = ?, description = ?, preconditions = ?, 
                test_steps = ?, expected_result = ?, priority = ?, test_type = ?, requirement_description = ?, status = ?
            WHERE id = ?
        ''', (
            data.get('test_case_name'),
            data.get('description'),
            data.get('preconditions'),
            data.get('test_steps'),
            data.get('expected_result'),
            data.get('priority'),
            data.get('test_type'),
            data.get('requirement_description'),
            data.get('status'),
            test_case_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Test case updated successfully'}), 200
    
    except Exception as e:
        print(f"Error updating test case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases/<int:test_case_id>', methods=['DELETE'])
def delete_test_case(test_case_id):
    """Delete a test case"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM test_cases WHERE id = ?', (test_case_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Test case deleted successfully'}), 200
    
    except Exception as e:
        print(f"Error deleting test case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/execute', methods=['POST'])
def execute_test_case():
    """Execute a test case for a specific version"""
    try:
        # Handle both JSON and FormData
        if request.is_json:
            data = request.json
        else:
            data = request.form
            
        test_case_id = data.get('test_case_id')
        status = data.get('status')
        version = data.get('version', 'Unspecified')
        comments = data.get('comments', '')
        defect_id = data.get('defect_id', '')
        
        # Handle file upload (Evidence)
        evidence_filename = None
        if 'evidence' in request.files:
            file = request.files['evidence']
            if file and file.filename != '':
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                safe_filename = secure_filename(f"evidence_{test_case_id}_{timestamp}_{file.filename}")
                evidence_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
                file.save(evidence_path)
                evidence_filename = safe_filename
        
        if not test_case_id or not status:
            return jsonify({'error': 'Missing test_case_id or status'}), 400
            
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # 1. Insert into history
        cursor.execute('''
            INSERT INTO test_executions (test_case_id, version, status, comments, defect_id, evidence_file)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (test_case_id, version, status, comments, defect_id, evidence_filename))
        
        # 2. Update master status
        cursor.execute('''
            UPDATE test_cases 
            SET status = ?
            WHERE id = ?
        ''', (status, test_case_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Execution recorded successfully', 'evidence': evidence_filename}), 200
        
    except Exception as e:
        print(f"Error executing test case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/reports', methods=['GET'])
def get_execution_reports():
    """Get aggregated execution statistics with optional version filtering"""
    try:
        version_filter = request.args.get('version', 'overall')
        
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Build query based on version filter
        if version_filter and version_filter.lower() != 'overall':
            # Filter by specific version
            cursor.execute('''
                SELECT version, status, COUNT(*) as count 
                FROM test_executions 
                WHERE version = ?
                GROUP BY version, status
                ORDER BY version
            ''', (version_filter,))
        else:
            # Get all versions (overall)
            cursor.execute('''
                SELECT version, status, COUNT(*) as count 
                FROM test_executions 
                GROUP BY version, status
                ORDER BY version
            ''')
        
        rows = cursor.fetchall()
        
        # Process data for charting
        # Structure: { "Release 1.0": { "Passed": 5, "Failed": 2, "Blocked": 0, "Total": 7 }, ... }
        report_data = {}
        
        for row in rows:
            version = row['version']
            status = row['status']
            count = row['count']
            
            if version not in report_data:
                report_data[version] = {'Passed': 0, 'Failed': 0, 'Blocked': 0, 'Total': 0}
            
            # Map status loosely (in case of 'Not Executed' etc)
            if status in report_data[version]:
                report_data[version][status] += count
            report_data[version]['Total'] += count
            
        conn.close()
        return jsonify(report_data), 200
        
    except Exception as e:
        print(f"Error generating reports: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/versions', methods=['GET'])
def get_versions():
    """Get list of unique versions from test executions"""
    try:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT DISTINCT version 
            FROM test_executions 
            ORDER BY executed_at DESC
        ''')
        rows = cursor.fetchall()
        
        versions = [row['version'] for row in rows]
        
        conn.close()
        return jsonify(versions), 200
        
    except Exception as e:
        print(f"Error fetching versions: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases-by-version', methods=['GET'])
def get_test_cases_by_version():
    """Get test cases with execution status for a specific version"""
    try:
        version = request.args.get('version', None)
        
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        if version and version.lower() != 'overall':
            # Get test cases with their execution status for specific version
            # Join with the most recent execution for this version
            cursor.execute('''
                SELECT 
                    tc.*,
                    COALESCE(te.status, 'Not Executed') as execution_status,
                    te.executed_at,
                    te.defect_id
                FROM test_cases tc
                LEFT JOIN (
                    SELECT test_case_id, status, executed_at, defect_id,
                           ROW_NUMBER() OVER (PARTITION BY test_case_id ORDER BY executed_at DESC) as rn
                    FROM test_executions
                    WHERE version = ?
                ) te ON tc.id = te.test_case_id AND te.rn = 1
                ORDER BY tc.created_at DESC
            ''', (version,))
        else:
            # Get all test cases with their latest status (any version)
            # We need to join to get the version of the latest execution
            cursor.execute('''
                SELECT 
                    tc.*,
                    COALESCE(te.status, 'Not Executed') as execution_status,
                    te.executed_at,
                    te.defect_id,
                    te.version
                FROM test_cases tc
                LEFT JOIN (
                    SELECT test_case_id, status, executed_at, defect_id, version,
                           ROW_NUMBER() OVER (PARTITION BY test_case_id ORDER BY executed_at DESC) as rn
                    FROM test_executions
                ) te ON tc.id = te.test_case_id AND te.rn = 1
                ORDER BY tc.created_at DESC
            ''')
        
        rows = cursor.fetchall()
        
        test_cases = []
        for row in rows:
            # For specific version filter, we know the version. For overall, we get it from the row.
            row_version = version if (version and version.lower() != 'overall') else (row['version'] if 'version' in row.keys() else None)
            
            test_cases.append({
                'id': row['id'],
                'requirement_file': row['requirement_file'],
                'requirement_id': row['requirement_id'] if 'requirement_id' in row.keys() else 'General',
                'requirement_description': row['requirement_description'] if 'requirement_description' in row.keys() else '',
                'test_case_name': row['test_case_name'],
                'description': row['description'],
                'preconditions': row['preconditions'],
                'test_steps': row['test_steps'],
                'expected_result': row['expected_result'],
                'priority': row['priority'],
                'test_type': row['test_type'],
                'status': row['execution_status'],
                'test_suite': row['test_suite'] if 'test_suite' in row.keys() else '',
                'created_at': row['created_at'],
                'executed_at': row['executed_at'] if 'executed_at' in row.keys() else None,
                'defect_id': row['defect_id'] if 'defect_id' in row.keys() else None,
                'version': row_version
            })
        
        conn.close()
        return jsonify(test_cases), 200
        
    except Exception as e:
        print(f"Error fetching test cases by version: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/defects', methods=['GET'])
def get_defects():
    """Get all failed test executions (Defects)"""
    try:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Join failures with test case details
        cursor.execute('''
            SELECT te.*, tc.test_case_name, tc.priority, tc.test_suite 
            FROM test_executions te 
            JOIN test_cases tc ON te.test_case_id = tc.id 
            WHERE te.status = 'Failed' 
            ORDER BY te.executed_at DESC
        ''')
        rows = cursor.fetchall()
        
        defects = []
        for row in rows:
            defects.append({
                'id': row['id'],
                'test_case_id': row['test_case_id'],
                'test_case_name': row['test_case_name'],
                'version': row['version'],
                'status': row['status'],
                'executed_at': row['executed_at'],
                'comments': row['comments'], # Contains the AI report
                'defect_id': row['defect_id'],
                'priority': row['priority'],
                'test_suite': row['test_suite']
            })
            
        conn.close()
        return jsonify(defects), 200
        
    except Exception as e:
        print(f"Error fetching defects: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/run-automation-script', methods=['POST'])
def run_automation_script():
    try:
        if not request.is_json:
            return jsonify({'error': 'Request must be JSON'}), 400
            
        data = request.json
        script_code = data.get('script')
        test_case_id = data.get('test_case_id')
        
        if not script_code:
            return jsonify({'error': 'No script provided'}), 400
            
        # Determine language (simple heuristic)
        extension = 'js'
        cmd = ['node']
        if 'import pytest' in script_code or 'from playwright' in script_code:
            extension = 'py'
            cmd = ['python'] # or pytest
            
        # Save to temp file
        filename = f"temp_test_{test_case_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{extension}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(script_code)
            
        # Execute
        import subprocess
        
        # Need to install playwright first? We assume environment is set up.
        # This is a dangerous operation in production, but okay for this local prototype.
        try:
            # Run with a timeout
            cmd.append(filepath)
            result = subprocess.run(
                cmd, 
                capture_output=True, 
                text=True, 
                timeout=30 # 30 seconds timeout
            )
            
            output = result.stdout + "\n" + result.stderr
            
            if result.returncode == 0:
                # Update DB status automatically?
                # For now let's just return success
                return jsonify({
                    'status': 'success', 
                    'output': output
                })
            else:
                return jsonify({
                    'status': 'failure', 
                    'error': 'Script execution failed',
                    'output': output
                })
                
        except subprocess.TimeoutExpired:
            return jsonify({'status': 'failure', 'error': 'Execution timed out (30s limit)', 'output': ''})
        except FileNotFoundError:
             return jsonify({'status': 'failure', 'error': f'Executable {cmd[0]} not found on server.', 'output': ''})

            
    except Exception as e:
        print(f"Error running script: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/generate-automation', methods=['POST'])
def generate_automation_script():
    try:
        print("--- Automation Request Received ---")
        # Check if we have a file or just JSON
        test_case_data = request.form.get('test_case')
        image_file = request.files.get('screenshot')
        
        print(f"Test Case Data Present: {bool(test_case_data)}")
        print(f"Image File Present: {bool(image_file)}")
        
        if not test_case_data:
            print("Error: No test case data")
            return jsonify({'error': 'No test case data provided'}), 400
            
        import json
        test_case = json.loads(test_case_data)
        target_url = request.form.get('target_url')
        target_username = request.form.get('target_username')
        target_password = request.form.get('target_password')
        
        print(f"Test Case Name: {test_case.get('test_case_name')}")
        print(f"Target URL: {target_url}")
        
        # Prepare the prompt
        prompt_text = f"""
        You are an expert QA Automation Engineer. 
        Generate a Playwright test script in JavaScript for the following test case.
        
        Test Case Name: {test_case.get('test_case_name')}
        Description: {test_case.get('description')}
        Preconditions: {test_case.get('preconditions')}
        Steps: {test_case.get('test_steps')}
        Expected Result: {test_case.get('expected_result')}
        Target URL: {target_url if target_url else "Not specified - Use generic 'http://localhost:3000'"}
        
        Rules:
        1. Use 'require("@playwright/test")' syntax.
        2. Use robust USER-FACING locators (getByRole, getByLabel, getByPlaceholder, getByText).
        3. AVOID using IDs or CSS selectors unless absolutely necessary.
        4. Include comments explaining the steps.
        5. Start the test by navigating to the Target URL: {target_url if target_url else "'http://localhost:3000'"}.
        6. Return ONLY the raw JavaScript code. Do not include markdown formatting.
        7. Include assertions based on the Expected Result.
        """
        
        if target_username and target_password:
             prompt_text += f"""
             8. IMPORTANT: The test involves logging in. 
                Use the following credentials if the test requires login:
                Username: "{target_username}"
                Password: "{target_password}"
                Use '.fill()' to enter these into appropriate fields detected by the context/screenshot.
             """
        
        # Use working model
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        if image_file:
            print(f"Processing screenshot: {image_file.filename}")
            # If screenshot provided, use it!
            from PIL import Image
            image = Image.open(image_file)
            print(f"Image opened successfully. Size: {image.size}")
            
            prompt_text += "\n\nI have attached a screenshot of the page. Use this visual information to determine the most accurate locators (button names, labels, placeholders)."
            
            print("Sending request with image to Gemini...")
            response = model.generate_content([prompt_text, image])
        else:
            print("Sending text-only request to Gemini...")
            # Text only generation
            response = model.generate_content(prompt_text)
        
        print("Response received from Gemini")
        # Clean up response
        script_content = response.text.replace('```javascript', '').replace('```', '').strip()
        print(f"Generated script length: {len(script_content)}")
        
        return jsonify({'script': script_content})

    except Exception as e:
        error_msg = str(e)
        print(f"Error generating automation script: {error_msg}")
        import traceback
        print(traceback.format_exc())
        
        # Check for quota exceeded error
        if "429" in error_msg or "quota" in error_msg.lower() or "resource_exhausted" in error_msg.lower():
            # Extract retry delay if available
            retry_seconds = 60  # Default
            if "retry_delay" in error_msg or "retry in" in error_msg.lower():
                import re
                # Try to extract seconds from error message
                match = re.search(r'retry in (\d+(?:\.\d+)?)', error_msg.lower())
                if match:
                    retry_seconds = int(float(match.group(1)))
            
            # Extract quota limit if available
            quota_limit = "unknown"
            if "limit:" in error_msg.lower():
                match = re.search(r'limit:\s*(\d+)', error_msg.lower())
                if match:
                    quota_limit = match.group(1)
            
            friendly_error = (
                f"⚠️ API Quota Exceeded for Automation Feature\n\n"
                f"FREE TIER LIMIT: {quota_limit} automation requests per day\n"
                f"RETRY IN: {retry_seconds} seconds ({retry_seconds//60} min {retry_seconds%60} sec)\n\n"
                f"SOLUTIONS:\n"
                f"1. ⏰ Wait {retry_seconds} seconds and try again\n"
                f"2. 🔑 Create a new API key: https://aistudio.google.com/app/apikey\n"
                f"3. 💳 Upgrade to paid tier: https://ai.google.dev/pricing\n\n"
                f"NOTE: Automation feature has separate quota from test case generation.\n"
                f"You can still generate test cases while waiting for automation quota to reset."
            )
            return jsonify({'error': friendly_error}), 429
        
        # For other errors, return generic message
        return jsonify({'error': f'Error generating automation script: {error_msg}'}), 500

@app.route('/test-ai', methods=['GET'])
def test_ai_connection():
    try:
        model = genai.GenerativeModel('gemini-1.5-flash')
        response = model.generate_content("Reply with 'API IS WORKING'")
        return jsonify({'status': 'success', 'message': response.text})
    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/test-cases/clear-all', methods=['DELETE'])
def clear_all_test_cases():
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM test_cases')
        deleted_count = cursor.rowcount
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': f'Successfully deleted {deleted_count} test cases'}), 200
    
    except Exception as e:
        print(f"Error clearing test cases: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export', methods=['GET'])
def export_test_cases():
    try:
        print("Export request received")
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM test_cases ORDER BY created_at DESC')
        rows = cursor.fetchall()
        
        conn.close()
        
        print(f"Found {len(rows)} test cases to export")
        
        if len(rows) == 0:
            return jsonify({'error': 'No test cases to export'}), 400
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define headers
        headers = ['ID', 'Test Suite / Module', 'Test Case Name', 'Description', 'Preconditions', 
                   'Test Steps', 'Expected Result', 'Priority', 'Test Type', 'Status',
                   'Requirement File', 'Created At']
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for row_num, row in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=row['id'])
            ws.cell(row=row_num, column=2, value=row['test_suite'] if 'test_suite' in row.keys() else '')
            ws.cell(row=row_num, column=3, value=row['test_case_name'])
            ws.cell(row=row_num, column=4, value=row['description'])
            ws.cell(row=row_num, column=5, value=row['preconditions'])
            ws.cell(row=row_num, column=6, value=row['test_steps'])
            ws.cell(row=row_num, column=7, value=row['expected_result'])
            ws.cell(row=row_num, column=8, value=row['priority'])
            ws.cell(row=row_num, column=9, value=row['test_type'])
            ws.cell(row=row_num, column=10, value=row['status'] if 'status' in row.keys() else 'Not Executed')
            ws.cell(row=row_num, column=11, value=row['requirement_file'])
            ws.cell(row=row_num, column=12, value=row['created_at'])
            
            # Apply text wrapping for long content
            for col in [3, 4, 5, 6, 7]:
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 20
        
        # Freeze the header row
        ws.freeze_panes = 'A2'

        # ==========================================
        # Create Traceability Matrix Sheet
        # ==========================================
        ws_rtm = wb.create_sheet(title="Traceability Matrix")
        
        # Define RTM headers
        rtm_headers = ['Requirement ID', 'Requirement Description', 'Test Case ID', 'Test Suite', 'Test Case Name', 'Status']
        
        # Write RTM headers
        for col_num, header in enumerate(rtm_headers, 1):
            cell = ws_rtm.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Group test cases by Requirement ID
        from collections import defaultdict
        req_map = defaultdict(list)
        for row in rows:
            # Handle potential missing requirement_id key if row structure varies (though we used select *)
            # Also handle if requirement_id is None
            r_id = row['requirement_id'] if row['requirement_id'] else 'General'
            req_map[r_id].append(row)
            
        # Sort requirements
        sorted_req_ids = sorted(req_map.keys(), key=lambda x: (x == 'General', x))

        current_row = 2
        
        # Styles for RTM
        center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        for req_id in sorted_req_ids:
            cases = req_map[req_id]
            start_row = current_row
            
            for tc in cases:
                # Requirement ID (will be merged later)
                ws_rtm.cell(row=current_row, column=1, value=req_id).alignment = center_align
                
                # Requirement Description (will be merged later)
                req_desc = tc['requirement_description'] if 'requirement_description' in tc.keys() and tc['requirement_description'] else 'No description'
                ws_rtm.cell(row=current_row, column=2, value=req_desc).alignment = left_align
                
                # Test Case ID
                ws_rtm.cell(row=current_row, column=3, value=f"TC-{tc['id']}").alignment = center_align

                # Test Suite
                ws_rtm.cell(row=current_row, column=4, value=tc['test_suite'] if 'test_suite' in tc.keys() else '').alignment = center_align
                
                # Test Case Name
                ws_rtm.cell(row=current_row, column=5, value=tc['test_case_name']).alignment = left_align
                
                # Status
                status = tc['status'] if 'status' in tc.keys() and tc['status'] else 'Not Executed'
                ws_rtm.cell(row=current_row, column=6, value=status).alignment = center_align
                
                current_row += 1
            
            end_row = current_row - 1
            
            # Merge Requirement ID and Description cells if there are multiple test cases
            if end_row > start_row:
                ws_rtm.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws_rtm.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

        # Adjust RTM column widths
        ws_rtm.column_dimensions['A'].width = 20
        ws_rtm.column_dimensions['B'].width = 50
        ws_rtm.column_dimensions['C'].width = 15
        ws_rtm.column_dimensions['D'].width = 20
        ws_rtm.column_dimensions['E'].width = 40
        ws_rtm.column_dimensions['F'].width = 15

        # Freeze RTM header
        ws_rtm.freeze_panes = 'A2'
        
        # Save to BytesIO instead of file
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        export_filename = f'test_cases_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        print(f"Sending file: {export_filename}")
        
        # Return file from memory
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=export_filename
        )
    
    except Exception as e:
        print(f"Error exporting test cases: {e}")
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/generate-defect', methods=['POST'])
def generate_defect_report():
    """Generate a structured bug report using AI"""
    try:
        data = request.json
        test_case = data.get('test_case', {})
        failure_reason = data.get('failure_reason', '')
        
        if not test_case:
            return jsonify({'error': 'Missing test case data'}), 400
            
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        prompt = f"""
        You are a QA Lead. A test case has FAILED. Your task is to write a PROFESSIONAL SOFTWARE BUG REPORT based on the details below.
        
        CONTEXT:
        Test Case Name: {test_case.get('test_case_name')}
        Test Steps: {test_case.get('test_steps')}
        Expected Result: {test_case.get('expected_result')}
        User's Observation (Failure Reason): "{failure_reason}"
        
        OUTPUT FORMAT (Markdown):
        **Summary:** [Concise, catchy title for the bug]
        
        **Description:**
        The user was executing the test "{test_case.get('test_case_name')}" and encountered an issue.
        
        **Steps to Reproduce:**
        [Refine the test steps into clear reproduction steps]
        
        **Actual Result:**
        [Based on the User's Observation]
        
        **Expected Result:**
        {test_case.get('expected_result')}
        
        **Severity:** [Suggest: Low/Medium/High/Critical based on context]
        **Environment:** [Suggest asking user, or assume generic]
        """
        
        
        response = model.generate_content(prompt)
        return jsonify({'report': response.text})
        
    except Exception as e:
        error_msg = str(e)
        print(f"Error generating defect report: {error_msg}")
        import traceback
        print(traceback.format_exc())
        
        # Check for quota exceeded error
        if "429" in error_msg or "quota" in error_msg.lower() or "resource_exhausted" in error_msg.lower():
            # Extract retry delay if available
            retry_seconds = 60  # Default
            if "retry_delay" in error_msg or "retry in" in error_msg.lower():
                import re
                match = re.search(r'retry in (\d+(?:\.\d+)?)', error_msg.lower())
                if match:
                    retry_seconds = int(float(match.group(1)))
            
            # Extract quota limit if available
            quota_limit = "unknown"
            if "limit:" in error_msg.lower():
                import re
                match = re.search(r'limit:\s*(\d+)', error_msg.lower())
                if match:
                    quota_limit = match.group(1)
            
            friendly_error = (
                f"⚠️ API Quota Exceeded for Defect Report Generation\n\n"
                f"FREE TIER LIMIT: {quota_limit} defect reports per day\n"
                f"RETRY IN: {retry_seconds} seconds ({retry_seconds//60} min {retry_seconds%60} sec)\n\n"
                f"SOLUTIONS:\n"
                f"1. ⏰ Wait {retry_seconds} seconds and try again\n"
                f"2. 🔑 Create a new API key: https://aistudio.google.com/app/apikey\n"
                f"3. 💳 Upgrade to paid tier: https://ai.google.dev/pricing\n"
                f"4. ✍️ Write the defect report manually (you can still submit the test failure)\n\n"
                f"NOTE: You can still mark the test as Failed and submit it without AI-generated report."
            )
            return jsonify({'error': friendly_error}), 429
        
        # For other errors, return generic message
        return jsonify({'error': f'Failed to generate defect report: {error_msg}'}), 500


def generate_test_cases_fallback(requirement_text, model_name='gemini-2.5-flash'):
    """Use Gemini AI to generate test cases with automatic fallback"""
    
    # Define fallback strategy with models that actually exist
    # Based on list_models.py output, only gemini-2.5-flash supports generateContent
    models_to_try = [model_name]
    
    # Add fallback models only if they're different from primary
    # Currently, gemini-2.5-flash is the only available model for text generation
    # Note: gemini-1.5-flash and gemini-1.5-pro are NOT available in current API
    if model_name != 'gemini-2.5-flash':
        models_to_try.append('gemini-2.5-flash')  # Fallback to the only working model
            
    last_error = None
    quota_exceeded_count = 0
    
    for current_model in models_to_try:
        try:
            print(f"Attempting generation with model: {current_model}")
            model = genai.GenerativeModel(current_model)
            
            prompt = f"""
            You are an expert QA engineer with 10+ years of experience. Analyze the following requirement document THOROUGHLY.
            First, identifying ALL distinct requirements and assign them IDs (e.g., REQ-001, REQ-002) if they are not already numbered.
            Then, generate COMPREHENSIVE test cases covering each of these requirements.
    
            REQUIREMENT DOCUMENT:
            {requirement_text[:30000]}
    
            CRITICAL INSTRUCTIONS:
            1. READ THE ENTIRE DOCUMENT carefully and identify EVERY feature, function, and requirement
            2. For EACH feature/requirement, generate test cases covering:
               - POSITIVE scenarios (happy path)
               - NEGATIVE scenarios (error handling)
               - EDGE CASES (boundary conditions)
               - INTEGRATION scenarios
    
            3. Generate detailed test cases in JSON format. Each test case MUST include:
               - test_case_name: A clear, specific, descriptive name
               - requirement_id: The ID of the requirement this test case covers (e.g., "REQ-001")
               - requirement_description: A brief description of the requirement itself (e.g. "User must log in")
               - description: Brief description of what is being tested and why
               - preconditions: Any setup or conditions needed before testing
               - test_steps: Detailed step-by-step instructions (as a numbered list with \\n separators)
               - expected_result: What should happen if the test passes
               - priority: High, Medium, or Low
               - test_type: Functional, Security, Performance, Usability, etc.
    
            TEST COVERAGE REQUIREMENTS:
            - Ensure EVERY identified requirement has at least one Positive and one Negative test case.
            - Maintain a 50-50 balance between Functional and Non-Functional tests where applicable.
    
            Return ONLY a valid JSON array of test cases.
            """
            
            response = model.generate_content(
                prompt,
                generation_config={"response_mime_type": "application/json"}
            )
            
            print(f"Received response from {current_model}")
            response_text = response.text.strip()
            
            # Clean up markdown code blocks if present
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.startswith('```'):
                response_text = response_text[3:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]
                
            response_text = response_text.strip()
            
            # Parse JSON response
            test_cases = json.loads(response_text)
            print(f"✅ Successfully parsed {len(test_cases)} test cases using {current_model}")
            return test_cases, None

        except Exception as e:
            error_msg = str(e)
            print(f"❌ Error with {current_model}: {error_msg}")
            
            # Check for quota exceeded (429 or "quota" in message)
            if "429" in error_msg or "quota" in error_msg.lower() or "resource_exhausted" in error_msg.lower():
                print(f"⚠️ Quota exceeded for {current_model}. Trying next model...")
                quota_exceeded_count += 1
                last_error = f"Quota exceeded for {current_model}"
                continue # Try next model
            else:
                # For non-quota errors, still try fallbacks
                print(f"⚠️ {current_model} failed: {error_msg}. Trying next model...")
                last_error = f"{current_model} error: {error_msg}"
                continue

    # If we exit the loop, all models failed
    if quota_exceeded_count == len(models_to_try):
        # All models hit quota limits
        return None, (
            f"⚠️ API Quota Exceeded - Rate limit reached for Gemini API.\n\n"
            f"FREE TIER LIMITS:\n"
            f"• 15 requests per minute (RPM)\n"
            f"• 1,500 requests per day (RPD)\n\n"
            f"IMMEDIATE SOLUTIONS:\n"
            f"1. ⏰ Wait 2-3 minutes and try again (quota resets every minute)\n"
            f"2. 🔑 Create a new API key at: https://aistudio.google.com/app/apikey\n"
            f"3. 💳 Upgrade to paid tier for 1,000 RPM: https://ai.google.dev/pricing\n\n"
            f"NOTE: Currently only 'gemini-2.5-flash' is available for text generation.\n"
            f"Other models (1.5-flash, 1.5-pro) are not available in this API version.\n\n"
            f"Run 'python quota_wait_timer.py' to wait with a countdown timer."
        )
    else:
        return None, f"All AI models failed. Last error: {last_error}. Please try again later."

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
