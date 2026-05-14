"""
Test the Excel export endpoint directly
"""
import requests

print("Testing Excel Export...")
print("=" * 60)

url = "http://localhost:5000/export"

print(f"Requesting: {url}")
response = requests.get(url)

print(f"Status Code: {response.status_code}")
print(f"Content-Type: {response.headers.get('Content-Type')}")
print(f"Content-Disposition: {response.headers.get('Content-Disposition')}")
print(f"Content-Length: {len(response.content)} bytes")

if response.status_code == 200:
    # Save the file
    filename = "test_export.xlsx"
    with open(filename, 'wb') as f:
        f.write(response.content)
    
    print(f"\n✅ SUCCESS!")
    print(f"File saved as: {filename}")
    print(f"File size: {len(response.content)} bytes")
    
    # Try to open with openpyxl to verify it's valid
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        print(f"Workbook loaded successfully!")
        print(f"Sheet name: {ws.title}")
        print(f"Rows: {ws.max_row}")
        print(f"Columns: {ws.max_column}")
        
        # Show first few rows
        print("\nFirst 3 rows:")
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            print(row[:3])  # Show first 3 columns
            
    except Exception as e:
        print(f"Error loading workbook: {e}")
else:
    print(f"\n✗ ERROR!")
    print(f"Response: {response.text}")
