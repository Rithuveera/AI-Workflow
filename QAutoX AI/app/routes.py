from flask import Blueprint, request, jsonify, render_template, send_file, current_app, url_for, session, redirect
from flask_login import login_required, current_user
import os
import json
import sqlite3
import uuid
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from .models import get_db_connection
from .services.file_service import extract_text_from_file, allowed_file
from .services.ai_service import (
    generate_test_cases, 
    generate_chat_response, 
    generate_defect_report_ai, 
    generate_automation_script_ai
)
from .services.confluence_service import fetch_confluence_page
from .services.api_test_service import (
    generate_api_test_cases_ai,
    parse_api_spec,
    execute_api_test,
    validate_response
)

main_bp = Blueprint('main', __name__)

# Global context for chat (simple in-memory storage)
current_context = ""

@main_bp.route('/')
@login_required
def index():
    if 'selected_product' not in session:
        return redirect(url_for('auth.select_product'))
    return render_template('index.html', user=current_user, selected_product=session['selected_product'])

@main_bp.route('/upload', methods=['POST'])
@login_required
def upload_file():
    """Handle file upload and generate test cases"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only PDF, DOCX, and TXT files are allowed'}), 400
    
    try:
        # Save file
        filename = secure_filename(file.filename)
        upload_folder = current_app.config['UPLOAD_FOLDER']
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)
        
        if 'selected_product' not in session:
            return jsonify({'error': 'No product selected. Please select a product first.'}), 400
            
        selected_product = session['selected_product']
        test_suite = request.form.get('test_suite', '')
        model_name = request.form.get('model', 'gemini-2.5-flash')
        
        # Extract text from file
        file_extension = filename.rsplit('.', 1)[1].lower()
        requirement_text = extract_text_from_file(file_path, file_extension)
        
        # Store requirement text for chat context
        global current_context
        current_context = requirement_text
        
        if not requirement_text:
            return jsonify({'error': 'Failed to extract text from file'}), 500
        
        # Generate test cases using AI Service
        print(f"Generating test cases for file: {filename} with model {model_name}")
        test_cases, error = generate_test_cases(requirement_text, model_name)
        
        if error:
            print(f"ERROR: {error}")
            return jsonify({'error': error}), 500
            
        if not test_cases:
            error_msg = 'Failed to generate test cases.'
            return jsonify({'error': error_msg}), 500
        
        # Save to DB
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Generat Batch ID
        batch_id = str(uuid.uuid4())
        
        # Save new test cases to database
        saved_test_cases = []
        for tc in test_cases:
            # Handle list fields if Gemini returns them as arrays
            test_steps = tc.get('test_steps', '')
            if isinstance(test_steps, list):
                test_steps = '\n'.join(test_steps)
                
            preconditions = tc.get('preconditions', '')
            if isinstance(preconditions, list):
                preconditions = '\n'.join(preconditions)

            cursor.execute('''
                INSERT INTO test_cases 
                (requirement_file, requirement_id, requirement_description, test_case_name, description, preconditions, test_steps, expected_result, priority, test_type, status, test_suite, product, batch_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                filename,
                tc.get('requirement_id', 'General'),
                tc.get('requirement_description', ''),
                tc.get('test_case_name', ''),
                tc.get('description', ''),
                preconditions,
                test_steps,
                tc.get('expected_result', ''),
                tc.get('priority', 'Medium'),
                tc.get('test_type', 'Functional'),
                'Not Executed',
                test_suite,
                selected_product,
                batch_id
            ))
            
            test_case_id = cursor.lastrowid
            saved_test_cases.append({
                'id': test_case_id,
                'requirement_file': filename,
                **tc
            })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Successfully generated {len(saved_test_cases)} test cases for {filename}',
            'filename': filename,
            'test_cases': saved_test_cases,
            'batch_id': batch_id
        }), 200
    
    except Exception as e:
        print(f"Error processing file: {e}")
        return jsonify({'error': str(e)}), 500

@main_bp.route('/import-confluence', methods=['POST'])
@login_required
def import_from_confluence():
    """Import content from Confluence and generate test cases"""
    try:
        data = request.json
        page_url = data.get('page_url')
        
        if not page_url:
            return jsonify({'error': 'No Confluence URL or Page ID provided'}), 400
            
        selected_product = session.get('selected_product')
        if not selected_product:
             return jsonify({'error': 'No product selected. Please select a product first.'}), 400

        test_suite = data.get('test_suite', '')
        model_name = data.get('model', 'gemini-2.5-flash')
        
        # Fetch from Confluence
        result, error = fetch_confluence_page(page_url)
        if error:
            return jsonify({'error': error}), 500
            
        requirement_text = result['content']
        page_title = result['title']
        filename = f"Confluence: {page_title}"
        
        # Store requirement text for chat context
        global current_context
        current_context = requirement_text
        
        # Generate test cases using AI Service
        print(f"Generating test cases for Confluence page: {page_title}")
        test_cases, error = generate_test_cases(requirement_text, model_name)
        
        if error:
            return jsonify({'error': error}), 500
            
        if not test_cases:
            return jsonify({'error': 'Failed to generate test cases.'}), 500
        
        # Save to DB
        conn = get_db_connection()
        cursor = conn.cursor()
        batch_id = str(uuid.uuid4())
        
        saved_test_cases = []
        for tc in test_cases:
            test_steps = tc.get('test_steps', '')
            if isinstance(test_steps, list):
                test_steps = '\n'.join(test_steps)
                
            preconditions = tc.get('preconditions', '')
            if isinstance(preconditions, list):
                preconditions = '\n'.join(preconditions)

            cursor.execute('''
                INSERT INTO test_cases 
                (requirement_file, requirement_id, requirement_description, test_case_name, description, preconditions, test_steps, expected_result, priority, test_type, status, test_suite, product, batch_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                filename,
                tc.get('requirement_id', 'General'),
                tc.get('requirement_description', ''),
                tc.get('test_case_name', ''),
                tc.get('description', ''),
                preconditions,
                test_steps,
                tc.get('expected_result', ''),
                tc.get('priority', 'Medium'),
                tc.get('test_type', 'Functional'),
                'Not Executed',
                test_suite,
                selected_product,
                batch_id
            ))
            
            test_case_id = cursor.lastrowid
            saved_test_cases.append({
                'id': test_case_id,
                'requirement_file': filename,
                **tc
            })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Successfully imported Confluence page and generated {len(saved_test_cases)} test cases',
            'filename': filename,
            'test_cases': saved_test_cases,
            'batch_id': batch_id
        }), 200
        
    except Exception as e:
        print(f"Confluence import error: {e}")
        return jsonify({'error': str(e)}), 500

@main_bp.route('/chat', methods=['POST'])
@login_required
def chat_with_ai():
    """Chat with the AI about the requirements"""
    try:
        data = request.json
        message = data.get('message')
        
        if not message:
            return jsonify({'error': 'No message provided'}), 400
            
        if not current_context:
            return jsonify({'reply': "I don't have any requirement document loaded yet. Please upload a file first."})
            
        reply = generate_chat_response(message, current_context)
        return jsonify({'reply': reply})
        
    except Exception as e:
        print(f"Chat error: {e}")
        return jsonify({'error': str(e)}), 500

@main_bp.route('/test-suites', methods=['GET'])
@login_required
def get_test_suites():
    """Get unique test suites (history)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        product = session.get('selected_product')
        
        cursor.execute('''
            SELECT batch_id, 
                   MAX(test_suite) as test_suite, 
                   MAX(requirement_file) as requirement_file, 
                   MAX(created_at) as created_at, 
                   COUNT(*) as total_cases
            FROM test_cases 
            WHERE product = ? AND batch_id IS NOT NULL
            GROUP BY batch_id
            ORDER BY created_at DESC
        ''', (product,))
        
        rows = cursor.fetchall()
        suites = []
        for row in rows:
            suites.append({
                'batch_id': row['batch_id'],
                'test_suite': row['test_suite'] or 'Unnamed Suite',
                'requirement_file': row['requirement_file'],
                'created_at': row['created_at'],
                'total_cases': row['total_cases']
            })
            
        conn.close()
        return jsonify(suites), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/test-cases', methods=['GET'])
@login_required
def get_test_cases():
    """Get all test cases or filter by filename"""
    try:
        filename = request.args.get('filename', None)
        conn = get_db_connection()
        cursor = conn.cursor()
        
        product = session.get('selected_product')
        
        if filename:
            if product:
                cursor.execute('SELECT * FROM test_cases WHERE requirement_file = ? AND product = ? ORDER BY created_at DESC', (filename, product))
            else:
                 cursor.execute('SELECT * FROM test_cases WHERE requirement_file = ? ORDER BY created_at DESC', (filename,))
        else:
            if product:
                cursor.execute('SELECT * FROM test_cases WHERE product = ? ORDER BY created_at DESC', (product,))
            else:
                cursor.execute('SELECT * FROM test_cases ORDER BY created_at DESC')
        rows = cursor.fetchall()
        
        test_cases = []
        for row in rows:
            req_id = row['requirement_id'] if 'requirement_id' in row.keys() else 'General'
            test_cases.append({
                'id': row['id'],
                'requirement_file': row['requirement_file'],
                'requirement_id': req_id,
                'requirement_description': row['requirement_description'] if 'requirement_description' in row.keys() else '',
                'test_case_name': row['test_case_name'],
                'description': row['description'],
                'preconditions': row['preconditions'],
                'test_steps': row['test_steps'],
                'expected_result': row['expected_result'],
                'priority': row['priority'],
                'test_type': row['test_type'],
                'status': row['status'] if 'status' in row.keys() else 'Not Executed',
                'test_suite': row['test_suite'] if 'test_suite' in row.keys() else '',
                'batch_id': row['batch_id'] if 'batch_id' in row.keys() else None,
                'created_at': row['created_at']
            })
        
        conn.close()
        return jsonify(test_cases), 200
    
    except Exception as e:
        print(f"Error fetching test cases: {e}")
        return jsonify({'error': str(e)}), 500

@main_bp.route('/test-cases/<int:test_case_id>', methods=['PUT'])
@login_required
def update_test_case(test_case_id):
    try:
        data = request.json
        conn = get_db_connection()
        conn.execute('''
            UPDATE test_cases 
            SET test_case_name = ?, description = ?, preconditions = ?, 
                test_steps = ?, expected_result = ?, priority = ?, test_type = ?, requirement_description = ?, status = ?
            WHERE id = ?
        ''', (
            data.get('test_case_name'),
            data.get('description'),
            data.get('preconditions'),
            data.get('test_steps'),
            data.get('expected_result'),
            data.get('priority'),
            data.get('test_type'),
            data.get('requirement_description'),
            data.get('status'),
            test_case_id
        ))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Test case updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/test-cases/<int:test_case_id>', methods=['DELETE'])
@login_required
def delete_test_case(test_case_id):
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM test_cases WHERE id = ?', (test_case_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Test case deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/execute', methods=['POST'])
@login_required
def execute_test_case():
    try:
        if request.is_json:
            data = request.json
        else:
            data = request.form
            
        test_case_id = data.get('test_case_id')
        status = data.get('status')
        version = data.get('version', 'Unspecified')
        comments = data.get('comments', '')
        defect_id = data.get('defect_id', '')
        
        evidence_filename = None
        if 'evidence' in request.files:
            file = request.files['evidence']
            if file and file.filename != '':
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                safe_filename = secure_filename(f"evidence_{test_case_id}_{timestamp}_{file.filename}")
                upload_folder = current_app.config['UPLOAD_FOLDER']
                evidence_path = os.path.join(upload_folder, safe_filename)
                file.save(evidence_path)
                evidence_filename = safe_filename
        
        if not test_case_id or not status:
            return jsonify({'error': 'Missing test_case_id or status'}), 400
            
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO test_executions (test_case_id, version, status, comments, defect_id, evidence_file)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (test_case_id, version, status, comments, defect_id, evidence_filename))
        
        conn.execute('UPDATE test_cases SET status = ? WHERE id = ?', (status, test_case_id))
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Execution recorded successfully', 'evidence': evidence_filename}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/reports', methods=['GET'])
@login_required
def get_execution_reports():
    try:
        version_filter = request.args.get('version', 'overall')
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if version_filter and version_filter.lower() != 'overall':
            cursor.execute('''
                SELECT version, status, COUNT(*) as count 
                FROM test_executions 
                WHERE version = ? AND test_case_id IN (SELECT id FROM test_cases WHERE product = ?)
                GROUP BY version, status
                ORDER BY version
            ''', (version_filter, session.get('selected_product')))
        else:
            cursor.execute('''
                SELECT version, status, COUNT(*) as count 
                FROM test_executions 
                WHERE test_case_id IN (SELECT id FROM test_cases WHERE product = ?)
                GROUP BY version, status
                ORDER BY version
            ''', (session.get('selected_product'),))
        
        rows = cursor.fetchall()
        report_data = {}
        
        for row in rows:
            version = row['version']
            status = row['status']
            count = row['count']
            
            if version not in report_data:
                report_data[version] = {'Passed': 0, 'Failed': 0, 'Blocked': 0, 'Total': 0}
            
            if status in report_data[version]:
                report_data[version][status] += count
            report_data[version]['Total'] += count
            
        conn.close()
        return jsonify(report_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/versions', methods=['GET'])
@login_required
def get_versions():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT DISTINCT te.version 
            FROM test_executions te
            JOIN test_cases tc ON te.test_case_id = tc.id
            WHERE tc.product = ?
            ORDER BY te.executed_at DESC
        ''', (session.get('selected_product'),))
        rows = cursor.fetchall()
        versions = [row['version'] for row in rows]
        conn.close()
        return jsonify(versions), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/test-cases-by-version', methods=['GET'])
@login_required
def get_test_cases_by_version():
    try:
        version = request.args.get('version', None)
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if version and version.lower() != 'overall':
            cursor.execute('''
                SELECT tc.*, COALESCE(te.status, 'Not Executed') as execution_status, te.executed_at, te.defect_id
                FROM test_cases tc
                LEFT JOIN (
                    SELECT test_case_id, status, executed_at, defect_id,
                           ROW_NUMBER() OVER (PARTITION BY test_case_id ORDER BY executed_at DESC) as rn
                    FROM test_executions
                    WHERE version = ?
                ) te ON tc.id = te.test_case_id AND te.rn = 1
                WHERE tc.product = ?
                ORDER BY tc.created_at DESC
            ''', (version, session.get('selected_product')))
        else:
            cursor.execute('''
                SELECT tc.*, COALESCE(te.status, 'Not Executed') as execution_status, te.executed_at, te.defect_id, te.version
                FROM test_cases tc
                LEFT JOIN (
                    SELECT test_case_id, status, executed_at, defect_id, version,
                           ROW_NUMBER() OVER (PARTITION BY test_case_id ORDER BY executed_at DESC) as rn
                    FROM test_executions
                ) te ON tc.id = te.test_case_id AND te.rn = 1
                WHERE tc.product = ?
                ORDER BY tc.created_at DESC
            ''', (session.get('selected_product'),))
        
        rows = cursor.fetchall()
        test_cases = []
        for row in rows:
            row_version = version if (version and version.lower() != 'overall') else (row['version'] if 'version' in row.keys() else None)
            test_cases.append({
                'id': row['id'],
                'requirement_file': row['requirement_file'],
                'requirement_id': row['requirement_id'] if 'requirement_id' in row.keys() else 'General',
                'requirement_description': row['requirement_description'] if 'requirement_description' in row.keys() else '',
                'test_case_name': row['test_case_name'],
                'description': row['description'],
                'preconditions': row['preconditions'],
                'test_steps': row['test_steps'],
                'expected_result': row['expected_result'],
                'priority': row['priority'],
                'test_type': row['test_type'],
                'status': row['execution_status'],
                'test_suite': row['test_suite'] if 'test_suite' in row.keys() else '',
                'created_at': row['created_at'],
                'executed_at': row['executed_at'] if 'executed_at' in row.keys() else None,
                'defect_id': row['defect_id'] if 'defect_id' in row.keys() else None,
                'version': row_version
            })
        
        conn.close()
        return jsonify(test_cases), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/defects', methods=['GET'])
@login_required
def get_defects():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT te.*, tc.test_case_name, tc.priority, tc.test_suite 
            FROM test_executions te 
            JOIN test_cases tc ON te.test_case_id = tc.id 
            WHERE te.status = 'Failed' AND tc.product = ?
            ORDER BY te.executed_at DESC
        ''', (session.get('selected_product'),))
        rows = cursor.fetchall()
        defects = []
        for row in rows:
            defects.append({
                'id': row['id'],
                'test_case_id': row['test_case_id'],
                'test_case_name': row['test_case_name'],
                'version': row['version'],
                'status': row['status'],
                'executed_at': row['executed_at'],
                'comments': row['comments'], 
                'defect_id': row['defect_id'],
                'priority': row['priority'],
                'test_suite': row['test_suite']
            })
        conn.close()
        return jsonify(defects), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/run-automation-script', methods=['POST'])
@login_required
def run_automation_script():
    try:
        if not request.is_json:
            return jsonify({'error': 'Request must be JSON'}), 400
        
        data = request.json
        script_code = data.get('script')
        test_case_id = data.get('test_case_id')
        
        if not script_code:
            return jsonify({'error': 'No script provided'}), 400
            
        extension = 'js'
        cmd = ['node']
        if 'import pytest' in script_code or 'from playwright' in script_code:
            extension = 'py'
            cmd = ['python']
            
        upload_folder = current_app.config['UPLOAD_FOLDER']
        filename = f"temp_test_{test_case_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{extension}"
        filepath = os.path.join(upload_folder, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(script_code)
            
        import subprocess
        try:
            cmd.append(filepath)
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            output = result.stdout + "\n" + result.stderr
            
            if result.returncode == 0:
                return jsonify({'status': 'success', 'output': output})
            else:
                return jsonify({'status': 'failure', 'error': 'Script execution failed', 'output': output})
        except subprocess.TimeoutExpired:
            return jsonify({'status': 'failure', 'error': 'Execution timed out', 'output': ''})
        except FileNotFoundError:
             return jsonify({'status': 'failure', 'error': f'Executable {cmd[0]} not found.', 'output': ''})
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/generate-automation', methods=['POST'])
@login_required
def generate_automation_route():
    try:
        test_case_data = request.form.get('test_case')
        image_file = request.files.get('screenshot')
        
        if not test_case_data:
            return jsonify({'error': 'No test case data provided'}), 400
            
        test_case = json.loads(test_case_data)
        target_url = request.form.get('target_url')
        target_username = request.form.get('target_username')
        target_password = request.form.get('target_password')
        
        script_content = generate_automation_script_ai(test_case, target_url, target_username, target_password, image_file)
        return jsonify({'script': script_content})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/test-ai', methods=['GET'])
def test_ai_connection():
    try:
        from .services.ai_service import configure_genai
        from google.generativeai import GenerativeModel
        configure_genai()
        model = GenerativeModel('gemini-2.5-flash') # Ensure checking correct model
        response = model.generate_content("Reply with 'API IS WORKING'")
        return jsonify({'status': 'success', 'message': response.text})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@main_bp.route('/test-cases/clear-all', methods=['DELETE'])
@login_required
def clear_all_test_cases():
    try:
        conn = get_db_connection()
        product = session.get('selected_product')
        if product:
             # Delete manual test cases
             conn.execute('DELETE FROM test_cases WHERE product = ?', (product,))
             
             # Delete executions for these manual test cases (optional but recommended)
             conn.execute('''
                DELETE FROM test_executions 
                WHERE test_case_id NOT IN (SELECT id FROM test_cases)
             ''')
             
             # Also clear API tests for this product systematically
             conn.execute('DELETE FROM api_test_cases WHERE product = ?', (product,))
             
             # Cleanup orphaned API executions
             conn.execute('''
                DELETE FROM api_test_executions 
                WHERE api_test_case_id NOT IN (SELECT id FROM api_test_cases)
             ''')
        else:
             return jsonify({'error': 'No product selected'}), 400
        conn.commit()
        conn.close()
        return jsonify({'message': 'Successfully deleted all test cases for selected product'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/api-testing/clear-all', methods=['DELETE'])
@login_required
def clear_all_api_tests():
    """Clear all API test cases for the selected product"""
    try:
        conn = get_db_connection()
        product = session.get('selected_product')
        if product:
            # Delete API test cases
            conn.execute('DELETE FROM api_test_cases WHERE product = ?', (product,))
            
            # Delete associated executions (cleanup orphaned records)
            conn.execute('''
                DELETE FROM api_test_executions 
                WHERE api_test_case_id NOT IN (SELECT id FROM api_test_cases)
            ''')
        else:
            return jsonify({'error': 'No product selected'}), 400
        conn.commit()
        conn.close()
        return jsonify({'message': 'Successfully deleted API test cases'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/export', methods=['GET'])
@login_required
def export_test_cases():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        product = session.get('selected_product')
        batch_id = request.args.get('batch_id')
        
        if product:
            if batch_id:
                cursor.execute('SELECT * FROM test_cases WHERE product = ? AND batch_id = ? ORDER BY created_at DESC', (product, batch_id))
            else:
                cursor.execute('SELECT * FROM test_cases WHERE product = ? ORDER BY created_at DESC', (product,))
        else:
            return jsonify({'error': 'No product selected'}), 400
        rows = cursor.fetchall()
        conn.close()
        
        if len(rows) == 0:
            return jsonify({'error': 'No test cases to export'}), 400
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        headers = ['ID', 'Test Suite / Module', 'Test Case Name', 'Description', 'Preconditions', 
                   'Test Steps', 'Expected Result', 'Priority', 'Test Type', 'Status',
                   'Requirement File', 'Created At']
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for row_num, row in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=row['id'])
            ws.cell(row=row_num, column=2, value=row['test_suite'] if 'test_suite' in row.keys() else '')
            ws.cell(row=row_num, column=3, value=row['test_case_name'])
            ws.cell(row=row_num, column=4, value=row['description'])
            ws.cell(row=row_num, column=5, value=row['preconditions'])
            ws.cell(row=row_num, column=6, value=row['test_steps'])
            ws.cell(row=row_num, column=7, value=row['expected_result'])
            ws.cell(row=row_num, column=8, value=row['priority'])
            ws.cell(row=row_num, column=9, value=row['test_type'])
            ws.cell(row=row_num, column=10, value=row['status'] if 'status' in row.keys() else 'Not Executed')
            ws.cell(row=row_num, column=11, value=row['requirement_file'])
            ws.cell(row=row_num, column=12, value=row['created_at'])
        
        # Traceability Matrix logic
        ws_rtm = wb.create_sheet(title="Traceability Matrix")
        rtm_headers = ['Requirement ID', 'Requirement Description', 'Test Case ID', 'Test Suite', 'Test Case Name', 'Status']
        for col_num, header in enumerate(rtm_headers, 1):
            cell = ws_rtm.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        from collections import defaultdict
        req_map = defaultdict(list)
        for row in rows:
            r_id = row['requirement_id'] if row['requirement_id'] else 'General'
            req_map[r_id].append(row)
        
        sorted_req_ids = sorted(req_map.keys(), key=lambda x: (x == 'General', x))
        current_row = 2
        for req_id in sorted_req_ids:
            cases = req_map[req_id]
            start_row = current_row
            for tc in cases:
                ws_rtm.cell(row=current_row, column=1, value=req_id)
                ws_rtm.cell(row=current_row, column=2, value=tc['requirement_description'] if 'requirement_description' in tc.keys() else '')
                ws_rtm.cell(row=current_row, column=3, value=f"TC-{tc['id']}")
                ws_rtm.cell(row=current_row, column=4, value=tc['test_suite'] if 'test_suite' in tc.keys() else '')
                ws_rtm.cell(row=current_row, column=5, value=tc['test_case_name'])
                ws_rtm.cell(row=current_row, column=6, value=tc['status'] if 'status' in tc.keys() else '')
                current_row += 1
            if current_row - 1 > start_row:
                 ws_rtm.merge_cells(start_row=start_row, start_column=1, end_row=current_row-1, end_column=1)
                 ws_rtm.merge_cells(start_row=start_row, start_column=2, end_row=current_row-1, end_column=2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        export_filename = f'test_cases_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/generate-defect', methods=['POST'])
@login_required
def generate_defect_report_route():
    try:
        data = request.json
        test_case = data.get('test_case', {})
        failure_reason = data.get('failure_reason', '')
        if not test_case:
            return jsonify({'error': 'Missing test case data'}), 400
            
        report = generate_defect_report_ai(test_case, failure_reason)
        return jsonify({'report': report})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/push-to-jira', methods=['POST'])
@login_required
def push_to_jira_route():
    from app.services.jira_service import create_jira_issue, attach_to_jira_issue
    try:
        data = request.form
        summary = data.get('summary')
        description = data.get('description')
        project_key = data.get('project_key')
        
        if not summary or not description:
            return jsonify({'error': 'Summary and Description are required'}), 400
            
        issue_key, error = create_jira_issue(summary, description, project_key)
        
        if error:
            return jsonify({'error': error}), 500
            
        # Handle attachment if present
        if 'evidence' in request.files:
            file = request.files['evidence']
            if file.filename != '':
                filename = secure_filename(file.filename)
                temp_path = os.path.join(current_app.config['UPLOAD_FOLDER'], f"jira_{issue_key}_{filename}")
                file.save(temp_path)
                
                success, attach_error = attach_to_jira_issue(issue_key, temp_path)
                # We don't fail the whole request if attachment fails, but we could log it
                if attach_error:
                    logger.error(f"Failed to attach evidence to Jira issue {issue_key}: {attach_error}")
                    
        return jsonify({
            'success': True,
            'issue_key': issue_key,
            'message': f'Jira ticket {issue_key} created successfully!'
        })
        
    except Exception as e:
        logger.exception("Error in push_to_jira_route")
        return jsonify({'error': str(e)}), 500

# --- API TESTING ROUTES ---

@main_bp.route('/api-testing/upload', methods=['POST'])
@login_required
def upload_api_spec():
    """Handle API specification upload and generate test cases"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        filename = secure_filename(file.filename)
        upload_folder = current_app.config['UPLOAD_FOLDER']
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)
        
        selected_product = session.get('selected_product')
        model_name = request.form.get('model', 'gemini-2.0-flash')
        
        # Parse spec
        spec_data, spec_type = parse_api_spec(file_path)
        if not spec_data:
            return jsonify({'error': 'Failed to parse API specification'}), 400
            
        # Generate test cases using AI
        spec_text = json.dumps(spec_data) if isinstance(spec_data, dict) else str(spec_data)
        test_cases, error = generate_api_test_cases_ai(spec_text, model_name)
        
        if error:
            return jsonify({'error': error}), 500
            
        # Save to DB
        conn = get_db_connection()
        batch_id = str(uuid.uuid4())
        saved_cases = []
        
        for tc in test_cases:
            cursor = conn.execute('''
                INSERT INTO api_test_cases 
                (api_name, method, endpoint, description, headers, query_params, request_body, 
                 expected_status_code, expected_response, assertions, test_type, priority, product, batch_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                tc.get('api_name'),
                tc.get('method'),
                tc.get('endpoint'),
                tc.get('description'),
                json.dumps(tc.get('headers', {})),
                json.dumps(tc.get('query_params', {})),
                json.dumps(tc.get('request_body', {})),
                tc.get('expected_status_code'),
                json.dumps(tc.get('expected_response', {})),
                json.dumps(tc.get('assertions', [])),
                tc.get('test_type', 'Functional'),
                tc.get('priority', 'Medium'),
                selected_product,
                batch_id
            ))
            
            tc_id = cursor.lastrowid
            saved_cases.append({
                'id': tc_id,
                **tc
            })
            
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Successfully generated {len(saved_cases)} API test cases',
            'test_cases': saved_cases,
            'batch_id': batch_id
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/api-testing/test-cases', methods=['GET'])
@login_required
def get_api_test_cases():
    """Get API test cases filtered by product or batch"""
    try:
        product = session.get('selected_product')
        batch_id = request.args.get('batch_id')
        
        conn = get_db_connection()
        if batch_id:
            rows = conn.execute('SELECT * FROM api_test_cases WHERE product = ? AND batch_id = ? ORDER BY created_at DESC', 
                               (product, batch_id)).fetchall()
        else:
            rows = conn.execute('SELECT * FROM api_test_cases WHERE product = ? ORDER BY created_at DESC', 
                               (product,)).fetchall()
        
        test_cases = []
        for row in rows:
            test_cases.append(dict(row))
            
        conn.close()
        return jsonify(test_cases), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/api-testing/execute/<int:test_id>', methods=['POST'])
@login_required
def run_api_test(test_id):
    """Execute a specific API test case"""
    try:
        conn = get_db_connection()
        test_case = conn.execute('SELECT * FROM api_test_cases WHERE id = ?', (test_id,)).fetchone()
        
        if not test_case:
            return jsonify({'error': 'Test case not found'}), 404
            
        # Execute request
        result, error = execute_api_test(
            method=test_case['method'],
            url=test_case['endpoint'],
            headers=test_case['headers'],
            params=test_case['query_params'],
            body=test_case['request_body']
        )
        
        if error:
            return jsonify({'error': error}), 500
            
        # Validate response
        validation_results = validate_response(
            result, 
            test_case['expected_status_code'],
            test_case['assertions']
        )
        
        overall_passed = all(r['passed'] for r in validation_results)
        status = 'Passed' if overall_passed else 'Failed'
        
        # Save execution
        version = request.json.get('version', 'Manual Run')
        conn.execute('''
            INSERT INTO api_test_executions 
            (api_test_case_id, version, status, response_time, status_code, response_body)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (test_id, version, status, result['response_time'], result['status_code'], result['body']))
        
        conn.execute('UPDATE api_test_cases SET status = ? WHERE id = ?', (status, test_id))
        conn.commit()
        conn.close()
        
        return jsonify({
            'status': status,
            'result': result,
            'validation': validation_results
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/api-testing/history/<int:test_id>', methods=['GET'])
@login_required
def get_api_history(test_id):
    """Get execution history for an API test case"""
    try:
        conn = get_db_connection()
        rows = conn.execute('SELECT * FROM api_test_executions WHERE api_test_case_id = ? ORDER BY executed_at DESC', 
                           (test_id,)).fetchall()
        
        history = [dict(row) for row in rows]
        conn.close()
        return jsonify(history), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
